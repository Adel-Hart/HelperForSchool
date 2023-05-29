# -*- coding: utf-8 -*- 
from openpyxl import Workbook
import os
import csv
import time
import keyboard

def cs():
    writeb = Workbook()
    writew = writeb.active
    writew['A1'] = "이름"
    writew['B1'] = "여부"
    dir = os.getcwd().replace("\\", "/").replace("/file", "")
    print(dir)
    tm = time.strftime('%Y-%m-%d', time.localtime(time.time()))
    if not os.path.isfile(dir+"/profile/profile.txt"):
        print("이름 파일이 없음")
        time.sleep(3)

    else:
        f = open(dir+"/profile/profile.txt", 'r', encoding="utf-8-sig")
        rang = int(f.readline())
        print(rang)
        with open(dir+"/data/dummy.dat", 'w', encoding="utf-8-sig") as dga:
            
            rangstr = rang
            for dg in range(rang):
                avc = f.readline().replace("\n", "")
                rangstr = int(rangstr)
                rangstr -= 1
                rangstr = str(rangstr)
                print(avc+"은 참여를 했나요? ("+rangstr+"명 남음)")
                cy = input("출석여부 :")
                if cy == "o":
                    dt = avc+" o\n"
                    dga.write(dt)
                else:
                    dt = avc+" x\n"
                    dga.write(dt)
                dg += 1
            dga.close()
            ad = 1
    print("done")
    
    with open(dir+"/data/dummy.dat", 'r', encoding="utf-8-sig") as ddir:
        for mt in range(rang):
            ddr = ddir.readline()
            acc = ddr.split(' ')[1]
            add = ddr.split(' ')[0]
            ad += 1
            writew.cell(ad,1,add)
            writew.cell(ad,2,acc)
        writeb.save(dir+'/exel/'+tm+".csv")
    dga.close()

    print("done")
    print("\nesc를 누르면 나감")
    while True:
        if keyboard.is_pressed('esc'):
            print("잘가요!")
            break